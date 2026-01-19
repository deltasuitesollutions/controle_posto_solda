from datetime import datetime
from io import BytesIO
from typing import Optional
from Server.services.export_service import buscar_registros, processar_linha

def exportar_registros_excel(
    data_inicio: Optional[str] = None, 
    data_fim: Optional[str] = None, 
    posto: Optional[str] = None
) -> BytesIO:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise Exception("Biblioteca openpyxl não instalada")
    
    rows = buscar_registros(data_inicio, data_fim, posto)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Registros de Produção"
    
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    center_alignment = Alignment(horizontal="center", vertical="center")
    
    headers = ['Posto', 'Inicio', 'Fim', 'Data', 'Produto', 'Matrícula', 'Operador']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_alignment
    
    column_widths = [12, 12, 12, 25, 25, 12, 25]  
    for col_num, width in enumerate(column_widths, 1):
        col_letter = ws.cell(row=1, column=col_num).column_letter
        ws.column_dimensions[col_letter].width = width
    
    for row_idx, row in enumerate(rows, 2):
        dados = processar_linha(row)
        
        for col_num in range(1, 8):
            cell = ws.cell(row=row_idx, column=col_num)
            
            if col_num == 1:
                cell.value = dados['posto'] or ''
            elif col_num == 2:
                cell.value = dados['hora_inicio'] or ''
            elif col_num == 3:
                cell.value = dados['hora_fim'] or ''
            elif col_num == 4:
                if dados['data_obj']:
                    # Usar objeto date do Python para o Excel reconhecer como data
                    cell.value = dados['data_obj']
                    # Formato de data padrão do Excel reconhecido universalmente
                    cell.number_format = 'dd/mm/yyyy'
                else:
                    cell.value = dados['data_str'] or ''
                    cell.number_format = 'General'
            elif col_num == 5:
                cell.value = dados['modelo_desc'] or ''
            elif col_num == 6:
                cell.value = dados['matricula'] or ''
            elif col_num == 7:
                cell.value = dados['nome'] or ''
            
            if col_num in [1, 2, 3, 4, 6]:
                cell.alignment = center_alignment
    
    ws.freeze_panes = 'A2'
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output

def gerar_nome_arquivo_excel() -> str:
    timestamp = datetime.now().strftime("%d%m%Y_%H%M%S")
    return f"registros_producao_{timestamp}.xlsx"
